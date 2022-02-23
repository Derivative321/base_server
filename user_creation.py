import openpyxl
import base64

from user import User

user = User()

user_database_file = openpyxl.load_workbook("user_data_base.xlsx")
user_database = user_database_file["Sheet1"]
data_base_next_row = user_database.max_row + 1


def get_and_validate_username(connection):
    flag = True
    while flag:
        username = connection.recv(2048).decode("utf-8")
        print(f"Incoming Username: {username}")
        if test_name_in_database(username):
            connection.sendall(str.encode("valid"))
            user_database.cell(data_base_next_row, 1).value = username
            user_database_file.save("user_data_base.xlsx")
            flag = False
        else:
            connection.sendall(str.encode("invalid"))


def test_name_in_database(username):
    user_list = {}
    for row in range(2, user_database.max_row + 1):
        usernames = user_database.cell(row, 1).value
        if usernames not in user_list:
            user_list[usernames] = 1

    for row in range(2, user_database.max_row + 1):
        if username not in user_list:
            return True
        else:
            return False


def get_and_validate_email(connection):
    flag = True
    while flag:
        email = connection.recv(2048).decode("utf-8")
        print(f"Incoming Email Address: {email}")
        if test_email_in_database(email):
            connection.sendall(str.encode("valid"))
            user_database.cell(data_base_next_row, 2).value = email
            user_database_file.save("user_data_base.xlsx")
            flag = False
        else:
            connection.sendall(str.encode("invalid"))


def test_email_in_database(email):
    email_list = {}
    for row in range(2, user_database.max_row + 1):
        emails = user_database.cell(row, 2).value
        if emails not in email_list:
            email_list[emails] = 1

    for row in range(2, user_database.max_row + 1):
        if email not in email_list:
            return True
        else:
            return False


def get_and_validate_password(connection):
    password_data = connection.recv(2048)
    encoded = base64.b64encode(password_data)
    user_database.cell(data_base_next_row, 3).value = encoded
    user_database_file.save("user_data_base.xlsx")
    print(base64.b64decode(encoded.decode("utf-8")))


def user_creation(connection):
    print("User Creation Initiated...")

    get_and_validate_username(connection)
    get_and_validate_email(connection)
    get_and_validate_password(connection)

    user_database.cell(data_base_next_row, 4).value = user.permissions
    user_database_file.save("user_data_base.xlsx")
